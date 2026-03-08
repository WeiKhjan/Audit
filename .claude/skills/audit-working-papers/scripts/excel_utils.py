"""
AUDIT EXCEL UTILITIES
=====================
Python utilities for creating Excel-based audit working papers and schedules.
Designed for Malaysian statutory audit engagements.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

# Colors
HEADER_BLUE = "1F4E79"
HEADER_LIGHT = "D6DCE5"
TOTAL_YELLOW = "FFFF00"
SUBTOTAL_GREY = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"

# Fonts
FONT_HEADER = Font(name='Arial', size=11, bold=True, color=WHITE)
FONT_TITLE = Font(name='Arial', size=14, bold=True)
FONT_SUBTITLE = Font(name='Arial', size=12, bold=True)
FONT_NORMAL = Font(name='Arial', size=10)
FONT_BOLD = Font(name='Arial', size=10, bold=True)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

THICK_BOTTOM = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

# Fills
FILL_HEADER = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type='solid')
FILL_TOTAL = PatternFill(start_color=TOTAL_YELLOW, end_color=TOTAL_YELLOW, fill_type='solid')
FILL_SUBTOTAL = PatternFill(start_color=SUBTOTAL_GREY, end_color=SUBTOTAL_GREY, fill_type='solid')
FILL_LIGHT = PatternFill(start_color=HEADER_LIGHT, end_color=HEADER_LIGHT, fill_type='solid')

# Alignment
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# ============================================================================
# WORKING PAPER TEMPLATES
# ============================================================================

def create_lead_schedule(
    filepath: str,
    client_name: str,
    year_end: str,
    subject: str,
    wp_ref: str,
    data: list,
    columns: list,
    preparer: str = "",
    reviewer: str = ""
):
    """
    Create a standard audit lead schedule.

    Args:
        filepath: Output file path
        client_name: Client company name
        year_end: Financial year end date
        subject: Working paper subject
        wp_ref: Working paper reference
        data: List of dictionaries containing schedule data
        columns: List of column headers
        preparer: Name of preparer
        reviewer: Name of reviewer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Schedule"

    # Header section
    ws.merge_cells('A1:H1')
    ws['A1'] = client_name
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:H2')
    ws['A2'] = f"YEAR END: {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    ws.merge_cells('A3:H3')
    ws['A3'] = subject
    ws['A3'].font = FONT_SUBTITLE

    # Working paper info
    ws['A5'] = "W/P Ref:"
    ws['B5'] = wp_ref
    ws['D5'] = "Prepared by:"
    ws['E5'] = preparer
    ws['F5'] = "Date:"
    ws['G5'] = datetime.now().strftime("%d/%m/%Y")

    ws['D6'] = "Reviewed by:"
    ws['E6'] = reviewer
    ws['F6'] = "Date:"
    ws['G6'] = ""

    # Column headers
    start_row = 8
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            if isinstance(value, (int, float)):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = ALIGN_LEFT

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 5, 12)

    wb.save(filepath)
    print(f"Lead schedule created: {filepath}")
    return filepath


def create_ppe_schedule(
    filepath: str,
    client_name: str,
    year_end: str,
    assets: list
):
    """
    Create Property, Plant & Equipment schedule with depreciation.

    Args:
        filepath: Output file path
        client_name: Client company name
        year_end: Financial year end date
        assets: List of asset dictionaries with keys:
            - description, date_acquired, cost_opening, additions, disposals, cost_closing
            - rate, accum_opening, depreciation, accum_disposal, accum_closing
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PPE Schedule"

    # Headers
    ws.merge_cells('A1:M1')
    ws['A1'] = f"CLIENT: {client_name}"
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:M2')
    ws['A2'] = f"SCHEDULE OF PROPERTY, PLANT & EQUIPMENT AS AT {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    # Column headers - Row 4
    headers_row1 = ['', '', 'COST', '', '', '', 'RATE', 'ACCUMULATED DEPRECIATION', '', '', '', 'CARRYING AMOUNT', '']
    headers_row2 = ['Date', 'Description', 'Opening', 'Additions', 'Disposals', 'Closing', '%', 'Opening', 'Charge', 'Disposals', 'Closing', 'Current', 'Prior']

    for col_idx, header in enumerate(headers_row2, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Data rows
    row = 5
    total_cost_open = total_cost_add = total_cost_disp = total_cost_close = 0
    total_accum_open = total_accum_charge = total_accum_disp = total_accum_close = 0
    total_nbv_curr = total_nbv_prior = 0

    for asset in assets:
        ws.cell(row=row, column=1, value=asset.get('date_acquired', '')).border = THIN_BORDER
        ws.cell(row=row, column=2, value=asset.get('description', '')).border = THIN_BORDER

        cost_open = asset.get('cost_opening', 0)
        cost_add = asset.get('additions', 0)
        cost_disp = asset.get('disposals', 0)
        cost_close = asset.get('cost_closing', cost_open + cost_add - cost_disp)

        rate = asset.get('rate', 0)

        accum_open = asset.get('accum_opening', 0)
        accum_charge = asset.get('depreciation', 0)
        accum_disp = asset.get('accum_disposal', 0)
        accum_close = asset.get('accum_closing', accum_open + accum_charge - accum_disp)

        nbv_curr = cost_close - accum_close
        nbv_prior = asset.get('nbv_prior', cost_open - accum_open)

        for col_idx, value in enumerate([cost_open, cost_add, cost_disp, cost_close], 3):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER

        ws.cell(row=row, column=7, value=f"{rate*100:.0f}%" if rate else "").border = THIN_BORDER

        for col_idx, value in enumerate([accum_open, accum_charge, accum_disp, accum_close], 8):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER

        for col_idx, value in enumerate([nbv_curr, nbv_prior], 12):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER

        # Accumulate totals
        total_cost_open += cost_open
        total_cost_add += cost_add
        total_cost_disp += cost_disp
        total_cost_close += cost_close
        total_accum_open += accum_open
        total_accum_charge += accum_charge
        total_accum_disp += accum_disp
        total_accum_close += accum_close
        total_nbv_curr += nbv_curr
        total_nbv_prior += nbv_prior

        row += 1

    # Total row
    ws.cell(row=row, column=2, value="TOTAL").font = FONT_BOLD
    for col_idx, value in enumerate([total_cost_open, total_cost_add, total_cost_disp, total_cost_close], 3):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.number_format = '#,##0.00'
        cell.font = FONT_BOLD
        cell.border = DOUBLE_BOTTOM

    for col_idx, value in enumerate([total_accum_open, total_accum_charge, total_accum_disp, total_accum_close], 8):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.number_format = '#,##0.00'
        cell.font = FONT_BOLD
        cell.border = DOUBLE_BOTTOM

    for col_idx, value in enumerate([total_nbv_curr, total_nbv_prior], 12):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.number_format = '#,##0.00'
        cell.font = FONT_BOLD
        cell.border = DOUBLE_BOTTOM

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    for col in 'CDEFGHIJKLM':
        ws.column_dimensions[col].width = 14

    wb.save(filepath)
    print(f"PPE schedule created: {filepath}")
    return filepath


def create_bank_reconciliation(
    filepath: str,
    client_name: str,
    year_end: str,
    bank_name: str,
    account_number: str,
    balance_per_bank: float,
    balance_per_books: float,
    unpresented_cheques: list,
    uncredited_deposits: list,
    other_reconciling_items: list = None
):
    """
    Create bank reconciliation working paper.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Reconciliation"

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = f"CLIENT: {client_name}"
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:E2')
    ws['A2'] = f"BANK RECONCILIATION AS AT {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    ws['A4'] = "Bank Name:"
    ws['B4'] = bank_name
    ws['A5'] = "Account No:"
    ws['B5'] = account_number

    # Reconciliation
    row = 7
    ws.cell(row=row, column=1, value="Balance per bank statement").font = FONT_BOLD
    ws.cell(row=row, column=5, value=balance_per_bank).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="Less: Unpresented cheques")
    row += 1

    total_unpresented = 0
    for item in unpresented_cheques:
        ws.cell(row=row, column=2, value=item.get('date', ''))
        ws.cell(row=row, column=3, value=item.get('cheque_no', ''))
        ws.cell(row=row, column=4, value=item.get('payee', ''))
        ws.cell(row=row, column=5, value=item.get('amount', 0)).number_format = '#,##0.00'
        total_unpresented += item.get('amount', 0)
        row += 1

    ws.cell(row=row, column=4, value="Total").font = FONT_BOLD
    ws.cell(row=row, column=5, value=total_unpresented).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="Add: Uncredited deposits")
    row += 1

    total_uncredited = 0
    for item in uncredited_deposits:
        ws.cell(row=row, column=2, value=item.get('date', ''))
        ws.cell(row=row, column=3, value=item.get('reference', ''))
        ws.cell(row=row, column=4, value=item.get('description', ''))
        ws.cell(row=row, column=5, value=item.get('amount', 0)).number_format = '#,##0.00'
        total_uncredited += item.get('amount', 0)
        row += 1

    ws.cell(row=row, column=4, value="Total").font = FONT_BOLD
    ws.cell(row=row, column=5, value=total_uncredited).number_format = '#,##0.00'

    row += 2
    adjusted_balance = balance_per_bank - total_unpresented + total_uncredited
    ws.cell(row=row, column=1, value="Adjusted bank balance").font = FONT_BOLD
    cell = ws.cell(row=row, column=5, value=adjusted_balance)
    cell.number_format = '#,##0.00'
    cell.font = FONT_BOLD
    cell.border = DOUBLE_BOTTOM

    row += 2
    ws.cell(row=row, column=1, value="Balance per books").font = FONT_BOLD
    cell = ws.cell(row=row, column=5, value=balance_per_books)
    cell.number_format = '#,##0.00'
    cell.font = FONT_BOLD
    cell.border = DOUBLE_BOTTOM

    row += 2
    difference = adjusted_balance - balance_per_books
    ws.cell(row=row, column=1, value="Difference").font = FONT_BOLD
    cell = ws.cell(row=row, column=5, value=difference)
    cell.number_format = '#,##0.00'
    cell.font = FONT_BOLD
    if abs(difference) > 0.01:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
        cell.font = Font(name='Arial', size=10, bold=True, color=WHITE)

    # Adjust widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15

    wb.save(filepath)
    print(f"Bank reconciliation created: {filepath}")
    return filepath


def create_trial_balance_template(
    filepath: str,
    client_name: str,
    year_end: str,
    accounts: list
):
    """
    Create trial balance template with proper formatting.

    Args:
        accounts: List of dicts with keys: acc_code, description, debit, credit
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    ws.merge_cells('A1:E1')
    ws['A1'] = client_name
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:E2')
    ws['A2'] = f"TRIAL BALANCE AS AT {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    headers = ['Acc Code', 'Description', 'Debit (RM)', 'Credit (RM)', 'W/P Ref']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    row = 5
    total_debit = total_credit = 0

    for acc in accounts:
        ws.cell(row=row, column=1, value=acc.get('acc_code', '')).border = THIN_BORDER
        ws.cell(row=row, column=2, value=acc.get('description', '')).border = THIN_BORDER

        debit = acc.get('debit', 0) or 0
        credit = acc.get('credit', 0) or 0

        cell_dr = ws.cell(row=row, column=3, value=debit if debit else '')
        cell_dr.number_format = '#,##0.00'
        cell_dr.border = THIN_BORDER

        cell_cr = ws.cell(row=row, column=4, value=credit if credit else '')
        cell_cr.number_format = '#,##0.00'
        cell_cr.border = THIN_BORDER

        ws.cell(row=row, column=5, value=acc.get('wp_ref', '')).border = THIN_BORDER

        total_debit += debit
        total_credit += credit
        row += 1

    # Totals
    ws.cell(row=row, column=2, value="TOTAL").font = FONT_BOLD
    cell_dr = ws.cell(row=row, column=3, value=total_debit)
    cell_dr.number_format = '#,##0.00'
    cell_dr.font = FONT_BOLD
    cell_dr.border = DOUBLE_BOTTOM

    cell_cr = ws.cell(row=row, column=4, value=total_credit)
    cell_cr.number_format = '#,##0.00'
    cell_cr.font = FONT_BOLD
    cell_cr.border = DOUBLE_BOTTOM

    # Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    wb.save(filepath)
    print(f"Trial balance created: {filepath}")
    return filepath


def create_pbc_checklist(
    filepath: str,
    client_name: str,
    year_end: str,
    items: list = None
):
    """
    Create PBC (Provided By Client) checklist.
    """
    if items is None:
        items = [
            # Statutory
            {"category": "STATUTORY DOCUMENTS", "item": "Certificate of Incorporation", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "Constitution / M&A", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "Form 24/49 (Current)", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "Board Resolutions", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "Minutes of AGM/EGM", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "SSM Company Profile", "required": "Yes"},
            {"category": "STATUTORY DOCUMENTS", "item": "Business Licenses", "required": "If applicable"},
            # Financial
            {"category": "FINANCIAL RECORDS", "item": "Trial Balance", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "General Ledger", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Bank Statements (Full Year)", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Bank Reconciliations", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Fixed Asset Register", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Debtor Aging", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Creditor Aging", "required": "Yes"},
            {"category": "FINANCIAL RECORDS", "item": "Inventory Listing", "required": "If applicable"},
            # Supporting
            {"category": "SUPPORTING DOCUMENTS", "item": "Sales Invoices (Sample)", "required": "Yes"},
            {"category": "SUPPORTING DOCUMENTS", "item": "Purchase Invoices (Sample)", "required": "Yes"},
            {"category": "SUPPORTING DOCUMENTS", "item": "Payment Vouchers (Sample)", "required": "Yes"},
            {"category": "SUPPORTING DOCUMENTS", "item": "Payroll Records", "required": "Yes"},
            {"category": "SUPPORTING DOCUMENTS", "item": "Contracts & Agreements", "required": "If applicable"},
            # Tax
            {"category": "TAX & COMPLIANCE", "item": "Tax Computation", "required": "Yes"},
            {"category": "TAX & COMPLIANCE", "item": "CP204 & Payment Receipts", "required": "Yes"},
            {"category": "TAX & COMPLIANCE", "item": "Form C (Prior Year)", "required": "Yes"},
            {"category": "TAX & COMPLIANCE", "item": "EPF Statements (Form A)", "required": "Yes"},
            {"category": "TAX & COMPLIANCE", "item": "SOCSO/EIS Statements", "required": "Yes"},
            # Confirmations
            {"category": "CONFIRMATIONS", "item": "Bank Confirmation", "required": "Auditor to send"},
            {"category": "CONFIRMATIONS", "item": "Debtor Confirmations", "required": "Sampling"},
            {"category": "CONFIRMATIONS", "item": "Creditor Confirmations", "required": "Sampling"},
            {"category": "CONFIRMATIONS", "item": "Related Party Confirmations", "required": "Yes"},
            {"category": "CONFIRMATIONS", "item": "Lawyer Confirmation", "required": "If applicable"},
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "PBC Checklist"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"CLIENT: {client_name}"
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:F2')
    ws['A2'] = f"PBC CHECKLIST - FYE {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    headers = ['No', 'Category', 'Document Required', 'Required', 'Received', 'Date Received', 'Remarks']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    row = 5
    current_category = ""
    item_no = 1

    for item in items:
        category = item.get('category', '')

        if category != current_category:
            # Category header row
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws.cell(row=row, column=1, value=category)
            cell.font = FONT_BOLD
            cell.fill = FILL_LIGHT
            current_category = category
            row += 1

        ws.cell(row=row, column=1, value=item_no).border = THIN_BORDER
        ws.cell(row=row, column=2, value="").border = THIN_BORDER
        ws.cell(row=row, column=3, value=item.get('item', '')).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item.get('required', '')).border = THIN_BORDER
        ws.cell(row=row, column=5, value="").border = THIN_BORDER  # Checkbox placeholder
        ws.cell(row=row, column=6, value="").border = THIN_BORDER
        ws.cell(row=row, column=7, value="").border = THIN_BORDER

        item_no += 1
        row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30

    wb.save(filepath)
    print(f"PBC checklist created: {filepath}")
    return filepath


def create_query_list(
    filepath: str,
    client_name: str,
    year_end: str,
    queries: list = None
):
    """
    Create audit queries tracking list.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Queries"

    ws.merge_cells('A1:H1')
    ws['A1'] = f"CLIENT: {client_name}"
    ws['A1'].font = FONT_TITLE

    ws.merge_cells('A2:H2')
    ws['A2'] = f"OUTSTANDING QUERIES LIST - FYE {year_end}"
    ws['A2'].font = FONT_SUBTITLE

    headers = ['No', 'W/P Ref', 'Query Description', 'Raised By', 'Date Raised', 'Client Response', 'Status', 'Closed Date']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    if queries:
        row = 5
        for idx, query in enumerate(queries, 1):
            ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
            ws.cell(row=row, column=2, value=query.get('wp_ref', '')).border = THIN_BORDER
            ws.cell(row=row, column=3, value=query.get('description', '')).border = THIN_BORDER
            ws.cell(row=row, column=4, value=query.get('raised_by', '')).border = THIN_BORDER
            ws.cell(row=row, column=5, value=query.get('date_raised', '')).border = THIN_BORDER
            ws.cell(row=row, column=6, value=query.get('response', '')).border = THIN_BORDER
            ws.cell(row=row, column=7, value=query.get('status', 'Open')).border = THIN_BORDER
            ws.cell(row=row, column=8, value=query.get('closed_date', '')).border = THIN_BORDER
            row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12

    wb.save(filepath)
    print(f"Query list created: {filepath}")
    return filepath


# ============================================================================
# DATA EXTRACTION UTILITIES
# ============================================================================

def read_trial_balance(filepath: str, sheet_name: str = None):
    """
    Read and parse trial balance from Excel file.
    Returns list of account dictionaries.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Try to identify account code and description columns
    accounts = []
    for idx, row in df.iterrows():
        # Skip empty rows and headers
        if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
            continue

        acc_code = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else ""

        # Look for account code patterns (e.g., 100-0000, 200-1000)
        if acc_code and '-' in acc_code:
            accounts.append({
                'acc_code': acc_code,
                'description': row.iloc[1] if len(row) > 1 else "",
                'debit': row.iloc[2] if len(row) > 2 and not pd.isna(row.iloc[2]) else 0,
                'credit': row.iloc[3] if len(row) > 3 and not pd.isna(row.iloc[3]) else 0
            })

    return accounts


def read_creditor_aging(filepath: str, sheet_name: str = None):
    """
    Read and parse creditor aging report.
    """
    df = pd.read_excel(filepath, sheet_name=sheet_name)
    return df


if __name__ == "__main__":
    # Test functions
    print("Audit Excel Utilities loaded successfully.")
    print("Available functions:")
    print("  - create_lead_schedule()")
    print("  - create_ppe_schedule()")
    print("  - create_bank_reconciliation()")
    print("  - create_trial_balance_template()")
    print("  - create_pbc_checklist()")
    print("  - create_query_list()")
    print("  - read_trial_balance()")
    print("  - read_creditor_aging()")
